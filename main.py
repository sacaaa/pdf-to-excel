import pytesseract
import cv2
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import Workbook
from os import path, listdir
from pdf2image import convert_from_path
from concurrent.futures import ThreadPoolExecutor
from numpy import array, float32
from datetime import datetime

POPPLER_PATH = "assets/poppler-23.11.0/Library/bin"
SIFT = cv2.SIFT_create()
BF = cv2.BFMatcher()

OUTPUT = "output_excel"
INPUT = "input_pdf"

ORIGINAL_IMAGE = cv2.imread('assets/original.jpg', 0)

WORKBOOK = Workbook()
SHEET = WORKBOOK.active

FONT_BOLD = Font(bold=True)
ALIGNMENT_CENTER = Alignment(horizontal='center')
FILL_YELLOW = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')

COORDINATES_TABLE = {
    "Date": {
        "position": [595, 366, 187, 39],
        "column": "A"
    },
    "Plate number": {
        "position": [87, 416, 236, 32],
        "column": "B"
    },
    "Delivery number": {
        "position": [967, 364, 211, 38],
        "column": "C"
    },
    "Material": {
        "position": [267, 609, 485, 30],
        "column": "D"
    },
    "Weight": {
        "position": [1049, 412, 129, 33],
        "column": "E"
    }
}

for key, value in COORDINATES_TABLE.items():
    value = value["column"] + "1"
    SHEET[value] = key
    SHEET[value].font = FONT_BOLD
    SHEET[value].alignment = ALIGNMENT_CENTER
    SHEET[value].fill = FILL_YELLOW


def is_pdf(file_path: str) -> bool:
    """Check if the file is a PDF file

    :param file_path: The path of the file
    :return: True if the file is a PDF file
    """

    return path.splitext(file_path)[1] == ".pdf"


def is_image(file_path: str) -> bool:
    """Check if the file is an image file

    :param file_path: The path of the file
    :return: True if the file is an image file
    """

    return path.splitext(file_path)[1] in [".jpg", ".jpeg", ".png", ".bmp"]


def convert_pdf_to_images(pdf_file):
    """Convert PDF file to images

    :param pdf_file: The PDF file
    :return: The images
    """

    return convert_from_path(path.join(INPUT, pdf_file), poppler_path=POPPLER_PATH)


def process_image(image):
    """Process the image to correct the perspective

    :param image: The image
    :return: The corrected image
    """

    # Detect keypoints and compute descriptors for the image and the original image
    kp1, des1 = SIFT.detectAndCompute(image, None)
    kp2, des2 = SIFT.detectAndCompute(ORIGINAL_IMAGE, None)

    # Match descriptors using Brute-Force Matcher
    matches = BF.knnMatch(des1, des2, k=2)

    # Filter out good matches using the ratio test
    good_matches = []
    for m, n in matches:
        if m.distance < 0.75 * n.distance:
            good_matches.append(m)

    # Compute the affine transformation matrix
    src_pts = float32([kp1[m.queryIdx].pt for m in good_matches]).reshape(-1, 1, 2)
    dst_pts = float32([kp2[m.trainIdx].pt for m in good_matches]).reshape(-1, 1, 2)
    affine_transformation_matrix, _ = cv2.estimateAffinePartial2D(src_pts, dst_pts)

    # Apply the affine transformation to the image
    rows, cols = image.shape
    corrected_image = cv2.warpAffine(image, affine_transformation_matrix, (cols, rows))

    return corrected_image


def process_pdf_file(pdf_file):
    """Process the PDF file to extract the data

    :param pdf_file: The PDF file
    :return: The processed images of the PDF file
    """

    images = convert_pdf_to_images(pdf_file)
    processed_images = []

    for i in range(len(images)):
        processed_images.append(process_image(cv2.cvtColor(array(images[i]), cv2.COLOR_BGR2GRAY)))

    return processed_images


def process_data_field(image, key: str, value: list, current_row: int):
    """Process a data field from the image

    :param image: The image
    :param key: The key of the data field
    :param value: The value of the data field
    :param current_row: The current row in the Excel sheet
    :return: The processed data
    """

    x, y, w, h = value["position"]
    cropped_image = image[y:y + h, x:x + w]

    data = pytesseract.image_to_string(cropped_image, lang="eng")
    processing_functions = {
        "Plate number": lambda data: data.split(",")[0],
        "Material": lambda data: data.split("(")[0].rstrip(),
        "Weight": lambda data: "".join(char for char in data if char.isdigit())
    }

    if key in processing_functions:
        data = processing_functions[key](data)

        if key == "Weight" and data != "":
            data = int(data)

    if data == "":
        print("Incorrect data detected! Excel row: " + str(2 + current_row))

    column_value = value["column"] + str(2 + current_row)
    SHEET[column_value] = data
    SHEET[column_value].alignment = ALIGNMENT_CENTER


def process_pdf_files(directory: str) -> None:
    """Process PDF files

    :param directory: The directory of the files
    """

    pdf_files = [file for file in listdir(directory) if is_pdf(file)]
    processed_pdfs = 0
    current_row = 0

    with ThreadPoolExecutor() as executor:
        results = list(executor.map(process_pdf_file, pdf_files))

    for pdf_file in results:
        for image in pdf_file:
            for key, value in COORDINATES_TABLE.items():
                process_data_field(image, key, value, current_row)

            current_row += 1
        processed_pdfs += 1

        WORKBOOK.save(path.join(OUTPUT, datetime.now().strftime("%Y-%m-%d_%H-%M") + ".xlsx"))
        print("Excel file saved!\nProcessed PDFs: " + str(processed_pdfs))


def check_existence() -> tuple[bool, str] | bool:
    """Check if the paths exist

    :return: True if all paths exist or a tuple with False and the path that does not exist
    """

    paths = [
        OUTPUT,
        INPUT,
        POPPLER_PATH,
        'assets/Tesseract-OCR/tesseract.exe',
        'assets/original.jpg'
    ]

    for path_name in paths:
        if not path.exists(path_name):
            return False, path_name

    if not is_image('assets/original.jpg'):
        return False, 'assets/original.jpg'

    return True


if __name__ == '__main__':
    existence_check = check_existence()

    if isinstance(existence_check, tuple):
        print(f"The path {existence_check[1]} does not exist.")
        exit()

    process_pdf_files(INPUT)

    print("Press any key to close the program...")

    from msvcrt import getch

    junk = getch()
